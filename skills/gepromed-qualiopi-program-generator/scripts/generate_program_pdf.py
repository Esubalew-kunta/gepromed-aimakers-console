#!/usr/bin/env python3
"""Generate a Qualiopi-compliant GEPROMED training PROGRAM as a print-ready,
brand-styled document (HTML — and a real .pdf when WeasyPrint is available).

This replaces the old .docx renderer. The deliverable is a **branded, A4,
print-ready HTML** document: open it and "Print -> Save as PDF" yields a
professional PDF with the GEPROMED charte (navy #007AC2 / #0A2540, orange
#ED6D1B accent, bundled logo, clean typography, `@page` / `@media print` CSS).
If the Python ``weasyprint`` library is importable, the script ALSO emits a real
``.pdf``; otherwise it emits the print-ready ``.html`` (graceful fallback, no
hard crash). Final Canva / InDesign polishing on a Gepromed gabarit is an
optional downstream design step — this HTML is the automated deliverable.

Inputs
------
1. **Schedule (required)** — an Excel workbook (``.xlsx``) read with ``openpyxl``.
   The schedule lives on a sheet named ``Planning`` (or the first sheet if that
   name is absent). Columns are EXACTLY, in this order::

       Jour | Heure début | Heure fin | Intitulé du créneau |
       Type (Cours / Atelier pratique) | Groupe (vide/"Tous" ou A, B...) |
       Salle | Encadrant(s) | Évalué (Oui/Non)

   One row = one time slot (créneau). Rows sharing the same day + start/end time
   but a different ``Groupe`` render as PARALLEL COLUMNS (one column per
   group/room). Rows whose ``Groupe`` is empty or "Tous" render as a single
   full-width slot.

2. **Fiche metadata (required)** — the Qualiopi header blocks (intitulé,
   référence, public visé, prérequis, objectifs, durée, modalités pédagogiques,
   moyens, évaluation, sanction, tarifs, inscription, contact, indicateurs…).
   Supplied EITHER by a first metadata sheet named ``Fiche`` in the SAME
   workbook (two columns: ``Champ`` / ``Valeur`` — one block per row; a value may
   repeat the ``Champ`` to append a bullet), OR by a companion JSON file passed
   with ``--meta``. The canonical, self-contained path is the ``Fiche`` sheet so
   one ``.xlsx`` carries everything; ``--meta`` overrides it when present.

Usage
-----
    # Render from a workbook that has both a "Fiche" and a "Planning" sheet
    python generate_program_pdf.py --in program.xlsx --out program.html

    # Schedule workbook + companion metadata JSON
    python generate_program_pdf.py --in schedule.xlsx --meta fiche.json --out program.pdf

    # Bundled demo (no input files needed) — sanity-check the styling
    python generate_program_pdf.py --demo --out demo.html

    # Write a ready-to-fill sample workbook (Fiche + Planning sheets)
    python generate_program_pdf.py --make-sample-xlsx sample_program.xlsx

    # Print the expected Excel column schema
    python generate_program_pdf.py --print-schema

If ``--out`` ends in ``.pdf`` and WeasyPrint is importable a real PDF is
written; otherwise a print-ready ``.html`` is written next to it and the script
still exits 0. Exit 2 on bad input.
"""
from __future__ import annotations

import argparse
import base64
import datetime as _dt
import html
import json
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
LOGO = SCRIPT_DIR.parent / "assets" / "gepromed-logo.png"

# --- GEPROMED brand constants -------------------------------------------------
NAVY = "#0A2540"        # deep navy — titles, dark surfaces
BLUE = "#007AC2"        # primary blue — headings, rules
ORANGE = "#ED6D1B"      # accent orange (rare, <=10%)
DARK = "#1F2A33"        # body text
MUTED = "#5F6B73"       # secondary text
TINT = "#E1F0F9"        # light blue tint (table headers / bands)
HAIRLINE = "#D9E2E8"

# The exact schedule columns (documented in references/excel-schedule-template.md).
SCHEDULE_COLUMNS = [
    "Jour",
    "Heure début",
    "Heure fin",
    "Intitulé du créneau",
    "Type (Cours / Atelier pratique)",
    'Groupe (vide/"Tous" ou A, B...)',
    "Salle",
    "Encadrant(s)",
    "Évalué (Oui/Non)",
]

# Fiche blocks in render order: (key, FR label, required)
BLOCKS = [
    ("public_vise", "Public visé", True),
    ("prerequis", "Prérequis", True),
    ("objectifs", "Objectifs pédagogiques", True),
    ("duree", "Durée", True),
    ("modalites_pedagogiques", "Modalités pédagogiques", True),
    ("moyens", "Moyens techniques et encadrement", False),
    ("evaluation", "Modalités d'évaluation", True),
    ("sanction", "Sanction / validation", False),
    ("accessibilite", "Accessibilité handicap", True),
    ("delais_acces", "Délais d'accès", True),
    ("tarifs", "Tarifs", True),
    ("inscription", "Modalités et délais d'inscription", True),
    ("contact", "Contact / référent pédagogique", False),
    ("indicateurs", "Indicateurs de résultats", False),
]

# Canonical FR field -> metadata key (accepts a few aliases for the Fiche sheet).
FICHE_FIELD_ALIASES = {
    "intitulé": "intitule",
    "intitule": "intitule",
    "référence": "reference",
    "reference": "reference",
    "version": "version",
    "date": "date",
    "public visé": "public_vise",
    "public vise": "public_vise",
    "prérequis": "prerequis",
    "prerequis": "prerequis",
    "objectifs": "objectifs",
    "objectifs pédagogiques": "objectifs",
    "durée": "duree",
    "duree": "duree",
    "modalités pédagogiques": "modalites_pedagogiques",
    "modalites pedagogiques": "modalites_pedagogiques",
    "moyens": "moyens",
    "moyens techniques et encadrement": "moyens",
    "évaluation": "evaluation",
    "evaluation": "evaluation",
    "modalités d'évaluation": "evaluation",
    "sanction": "sanction",
    "accessibilité": "accessibilite",
    "accessibilite": "accessibilite",
    "accessibilité handicap": "accessibilite",
    "délais d'accès": "delais_acces",
    "delais d'acces": "delais_acces",
    "delais acces": "delais_acces",
    "tarifs": "tarifs",
    "tarif": "tarifs",
    "inscription": "inscription",
    "modalités et délais d'inscription": "inscription",
    "contact": "contact",
    "référent pédagogique": "contact",
    "indicateurs": "indicateurs",
    "indicateurs de résultats": "indicateurs",
}

DEFAULT_ACCESSIBILITE = (
    "GEPROMED s'engage à étudier toute situation de handicap afin d'envisager "
    "l'adaptation de la formation et des modalités d'accueil. Un référent handicap "
    "est à votre disposition pour analyser votre demande au cas par cas et "
    "déterminer les aménagements possibles. Contact référent handicap : "
    "[email / téléphone à confirmer]."
)
DEFAULT_SANCTION = (
    "Attestation de fin de formation remise à chaque participant. "
    "Certificat de réalisation établi pour les actions concernées."
)

PLACEHOLDER = "[À compléter — valider par le RQ]"


# ---------------------------------------------------------------------------
# Value / cell helpers
# ---------------------------------------------------------------------------
def _cell_str(value) -> str:
    """Normalise an openpyxl cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, _dt.time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, _dt.datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_group(value: str) -> str:
    v = (value or "").strip()
    if v == "" or v.lower() in ("tous", "toutes", "all", "commun"):
        return "Tous"
    return v


def _split_multi(value: str) -> list[str]:
    """Split a metadata value into bullets on newlines or ' | ' separators."""
    if value is None:
        return []
    parts = []
    for chunk in str(value).replace("\r", "\n").split("\n"):
        for sub in chunk.split(" | "):
            s = sub.strip(" -•\t")
            if s:
                parts.append(s)
    return parts


# ---------------------------------------------------------------------------
# Excel reading (openpyxl)
# ---------------------------------------------------------------------------
def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError as e:  # pragma: no cover
        raise SystemExit(
            "ERROR: this generator needs the 'openpyxl' package to read .xlsx.\n"
            "Install it with:  pip install openpyxl"
        ) from e


def read_workbook(path: Path) -> tuple[dict, list[dict]]:
    """Return (metadata, slots) read from an .xlsx workbook.

    - metadata: from the 'Fiche' sheet if present (else {}).
    - slots: from the 'Planning' sheet if present, else the first non-Fiche sheet.
    """
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheet_names = wb.sheetnames

    metadata: dict = {}
    fiche_name = next((n for n in sheet_names if n.strip().lower() == "fiche"), None)
    if fiche_name:
        metadata = _read_fiche_sheet(wb[fiche_name])

    planning_name = next(
        (n for n in sheet_names if n.strip().lower() in ("planning", "programme", "schedule")),
        None,
    )
    if planning_name is None:
        planning_name = next((n for n in sheet_names if n != fiche_name), None)
    if planning_name is None:
        raise SystemExit("ERROR: workbook has no schedule sheet (expected 'Planning').")

    slots = _read_planning_sheet(wb[planning_name])
    wb.close()
    return metadata, slots


def _read_fiche_sheet(ws) -> dict:
    """Two-column key/value sheet. Repeated keys append bullets into a list."""
    meta: dict = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        field = _cell_str(row[0]) if len(row) > 0 else ""
        value = _cell_str(row[1]) if len(row) > 1 else ""
        if not field:
            continue
        low = field.strip().lower().rstrip(":")
        if low in ("champ", "field", "clé", "cle", "key"):
            continue  # header row
        key = FICHE_FIELD_ALIASES.get(low)
        if key is None:
            continue
        if not value:
            continue
        if key in meta:
            existing = meta[key]
            if isinstance(existing, list):
                existing.append(value)
            else:
                meta[key] = [existing, value]
        else:
            # Values holding embedded separators become a bullet list.
            multi = _split_multi(value)
            meta[key] = multi if len(multi) > 1 else value
    return meta


def _read_planning_sheet(ws) -> list[dict]:
    """Read schedule rows keyed by the documented column order.

    Matches on the header row (case/space-insensitive) so column order in the
    file is tolerated; falls back to positional mapping if no header matches.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    def norm(s: str) -> str:
        return "".join(ch for ch in _cell_str(s).lower() if ch.isalnum())

    canonical = {
        norm("Jour"): "jour",
        norm("Heure début"): "debut",
        norm("Heure fin"): "fin",
        norm("Intitulé du créneau"): "intitule",
        norm("Type (Cours / Atelier pratique)"): "type",
        norm('Groupe (vide/"Tous" ou A, B...)'): "groupe",
        norm("Type"): "type",
        norm("Groupe"): "groupe",
        norm("Salle"): "salle",
        norm("Encadrant(s)"): "encadrants",
        norm("Encadrant"): "encadrants",
        norm("Évalué (Oui/Non)"): "evalue",
        norm("Évalué"): "evalue",
        norm("Evalue"): "evalue",
    }

    header = rows[0]
    col_map: dict[int, str] = {}
    matched = 0
    for i, cell in enumerate(header):
        key = canonical.get(norm(cell))
        if key:
            col_map[i] = key
            matched += 1

    if matched >= 3:
        data_rows = rows[1:]
    else:
        # No usable header — assume positional order per SCHEDULE_COLUMNS.
        positional = ["jour", "debut", "fin", "intitule", "type", "groupe",
                      "salle", "encadrants", "evalue"]
        col_map = {i: positional[i] for i in range(min(len(positional), len(header)))}
        data_rows = rows

    slots: list[dict] = []
    for r in data_rows:
        if r is None:
            continue
        rec = {"jour": "", "debut": "", "fin": "", "intitule": "", "type": "",
               "groupe": "", "salle": "", "encadrants": "", "evalue": ""}
        for i, key in col_map.items():
            if i < len(r):
                rec[key] = _cell_str(r[i])
        if not any(rec[k] for k in ("jour", "debut", "intitule")):
            continue  # skip blank rows
        slots.append(rec)
    return slots


# ---------------------------------------------------------------------------
# Timetable grouping — parallel sub-group columns
# ---------------------------------------------------------------------------
def build_timetable(slots: list[dict]) -> list[dict]:
    """Group slots by (Jour) then by (Heure début, Heure fin).

    Returns an ordered list of days; each day has time blocks; each block carries
    the 'Tous' (full-width) rows and a per-group column map for concurrent rows.
    """
    days: list[dict] = []
    day_index: dict[str, dict] = {}

    for s in slots:
        jour = s["jour"] or "Jour"
        if jour not in day_index:
            entry = {"jour": jour, "_blocks": {}, "_order": []}
            day_index[jour] = entry
            days.append(entry)
        day = day_index[jour]
        tkey = (s["debut"], s["fin"])
        if tkey not in day["_blocks"]:
            block = {"debut": s["debut"], "fin": s["fin"], "common": [], "groups": {}}
            day["_blocks"][tkey] = block
            day["_order"].append(tkey)
        block = day["_blocks"][tkey]
        g = _norm_group(s["groupe"])
        if g == "Tous":
            block["common"].append(s)
        else:
            block["groups"].setdefault(g, []).append(s)

    # Materialise ordered structure and sort time blocks by start time.
    out: list[dict] = []
    for day in days:
        blocks = []
        for tkey in sorted(day["_order"], key=lambda k: (k[0] or "", k[1] or "")):
            b = day["_blocks"][tkey]
            b["group_names"] = sorted(b["groups"].keys())
            b["parallel"] = len(b["group_names"]) > 0
            blocks.append(b)
        out.append({"jour": day["jour"], "blocks": blocks})
    return out


# ---------------------------------------------------------------------------
# HTML rendering
# ---------------------------------------------------------------------------
def _esc(s) -> str:
    return html.escape(str(s), quote=True)


def _logo_data_uri() -> str | None:
    if not LOGO.exists():
        return None
    try:
        b = LOGO.read_bytes()
        return "data:image/png;base64," + base64.b64encode(b).decode("ascii")
    except Exception:  # noqa: BLE001
        return None


def _render_block_value(value) -> str:
    if value is None or value == "" or value == []:
        return f'<p class="ph">{_esc(PLACEHOLDER)}</p>'
    if isinstance(value, list):
        items = "".join(f"<li>{_esc(v)}</li>" for v in value if str(v).strip())
        return f"<ul>{items}</ul>" if items else f'<p class="ph">{_esc(PLACEHOLDER)}</p>'
    return f"<p>{_esc(value)}</p>"


def _slot_card_html(s: dict) -> str:
    typ = s.get("type", "")
    is_atelier = "atelier" in typ.lower()
    kind_cls = "atelier" if is_atelier else "cours"
    evalue = s.get("evalue", "").strip().lower() in ("oui", "yes", "o", "x", "true", "1")
    meta_bits = []
    if s.get("salle"):
        meta_bits.append(f'<span class="meta"><b>Salle :</b> {_esc(s["salle"])}</span>')
    if s.get("encadrants"):
        meta_bits.append(f'<span class="meta"><b>Encadrant(s) :</b> {_esc(s["encadrants"])}</span>')
    if typ:
        meta_bits.append(f'<span class="tag {kind_cls}">{_esc(typ)}</span>')
    if evalue:
        meta_bits.append('<span class="tag evalue">Évalué</span>')
    meta = " ".join(meta_bits)
    return (
        f'<div class="slot {kind_cls}">'
        f'<div class="slot-title">{_esc(s.get("intitule") or "—")}</div>'
        f'<div class="slot-meta">{meta}</div>'
        f"</div>"
    )


def _timeblock_html(block: dict) -> str:
    time_lbl = _esc(block["debut"])
    if block["fin"]:
        time_lbl += " – " + _esc(block["fin"])
    parts = [f'<div class="trow">',
             f'<div class="tcell time"><span>{time_lbl}</span></div>',
             f'<div class="tcell body">']

    # Full-width 'Tous' rows first.
    for s in block["common"]:
        parts.append(_slot_card_html(s))

    # Parallel group columns.
    if block["parallel"]:
        cols = []
        for g in block["group_names"]:
            slot_html = "".join(_slot_card_html(s) for s in block["groups"][g])
            cols.append(
                f'<div class="gcol"><div class="ghead">Groupe {_esc(g)}</div>{slot_html}</div>'
            )
        parts.append(f'<div class="parallel cols-{len(cols)}">{"".join(cols)}</div>')

    parts.append("</div></div>")
    return "".join(parts)


def _timetable_html(timetable: list[dict]) -> str:
    if not timetable:
        return '<p class="ph">[Planning à compléter — importer un fichier Excel]</p>'
    out = ['<section class="block"><h2>Contenu / planning détaillé</h2>']
    for day in timetable:
        out.append(f'<div class="day"><div class="day-head">{_esc(day["jour"])}</div>')
        for block in day["blocks"]:
            out.append(_timeblock_html(block))
        out.append("</div>")
    out.append("</section>")
    return "".join(out)


CSS = f"""
:root {{
  --navy: {NAVY}; --blue: {BLUE}; --orange: {ORANGE};
  --dark: {DARK}; --muted: {MUTED}; --tint: {TINT}; --hair: {HAIRLINE};
}}
* {{ box-sizing: border-box; }}
html, body {{ margin: 0; padding: 0; }}
body {{
  font-family: "Segoe UI", "Helvetica Neue", Arial, sans-serif;
  color: var(--dark); font-size: 10.8pt; line-height: 1.5;
  background: #f4f6f8;
}}
.sheet {{
  background: #fff; max-width: 210mm; margin: 12px auto; padding: 16mm 18mm;
  box-shadow: 0 1px 6px rgba(10,37,64,.12);
}}
header.masthead {{ display: flex; align-items: flex-start; justify-content: space-between;
  gap: 16px; border-bottom: 3px solid var(--blue); padding-bottom: 12px; }}
header.masthead img.logo {{ height: 46px; }}
header.masthead .brandline {{ font-size: 8.2pt; color: var(--muted); font-style: italic;
  margin-top: 4px; }}
.doc-tag {{ display: inline-block; background: var(--orange); color: #fff; font-weight: 700;
  font-size: 8.5pt; letter-spacing: .04em; text-transform: uppercase;
  padding: 3px 9px; border-radius: 3px; }}
h1.title {{ color: var(--navy); font-size: 20pt; line-height: 1.15; margin: 14px 0 4px; }}
.refline {{ color: var(--muted); font-size: 8.6pt; margin: 0 0 6px; }}
.block {{ margin-top: 16px; break-inside: avoid; }}
.block h2 {{ color: var(--blue); font-size: 12.5pt; margin: 0 0 6px;
  border-bottom: 1px solid var(--blue); padding-bottom: 3px; }}
.block p {{ margin: 3px 0; }}
.block ul {{ margin: 4px 0 4px 18px; padding: 0; }}
.block li {{ margin: 2px 0; }}
.ph {{ color: var(--muted); font-style: italic; }}
/* Timetable */
.day {{ margin-top: 12px; break-inside: avoid; }}
.day-head {{ background: var(--navy); color: #fff; font-weight: 700; font-size: 10.5pt;
  padding: 5px 10px; border-radius: 3px 3px 0 0; }}
.trow {{ display: flex; border: 1px solid var(--hair); border-top: none; }}
.tcell.time {{ flex: 0 0 78px; background: var(--tint); color: var(--navy); font-weight: 700;
  font-size: 8.8pt; padding: 8px; display: flex; align-items: flex-start; }}
.tcell.body {{ flex: 1 1 auto; padding: 8px; min-width: 0; }}
.slot {{ border-left: 3px solid var(--blue); background: #fbfdff; padding: 6px 9px;
  margin: 0 0 6px; border-radius: 0 3px 3px 0; }}
.slot:last-child {{ margin-bottom: 0; }}
.slot.atelier {{ border-left-color: var(--orange); background: #fff7f0; }}
.slot-title {{ font-weight: 600; color: var(--dark); font-size: 10pt; }}
.slot-meta {{ margin-top: 3px; font-size: 8.3pt; color: var(--muted); }}
.slot-meta .meta {{ margin-right: 10px; }}
.tag {{ display: inline-block; font-size: 7.6pt; font-weight: 700; text-transform: uppercase;
  letter-spacing: .03em; padding: 1px 6px; border-radius: 8px; margin-right: 5px; }}
.tag.cours {{ background: var(--tint); color: var(--blue); }}
.tag.atelier {{ background: #fde6d3; color: var(--orange); }}
.tag.evalue {{ background: #e7f6ec; color: #1a7f43; }}
.parallel {{ display: flex; gap: 8px; }}
.gcol {{ flex: 1 1 0; min-width: 0; border: 1px solid var(--hair); border-radius: 3px;
  padding: 6px; background: #fff; }}
.ghead {{ font-size: 8.4pt; font-weight: 700; color: var(--navy); text-transform: uppercase;
  letter-spacing: .04em; margin-bottom: 5px; border-bottom: 1px dashed var(--hair);
  padding-bottom: 3px; }}
.compliance {{ margin-top: 18px; padding: 9px 12px; background: var(--tint);
  border-radius: 4px; font-size: 8.4pt; color: var(--muted); font-style: italic; }}
.printbar {{ position: sticky; top: 0; z-index: 10; background: var(--navy); color: #fff;
  display: flex; align-items: center; justify-content: space-between; gap: 12px;
  padding: 8px 16px; font-size: 9.5pt; }}
.printbar button {{ background: var(--orange); color: #fff; border: 0; font-weight: 700;
  padding: 7px 16px; border-radius: 4px; cursor: pointer; font-size: 9.5pt; }}
.footer {{ margin-top: 16px; border-top: 1px solid var(--hair); padding-top: 6px;
  font-size: 7.6pt; color: var(--muted); text-align: center; }}
@page {{ size: A4; margin: 14mm 15mm; }}
@media print {{
  body {{ background: #fff; }}
  .sheet {{ box-shadow: none; margin: 0; max-width: none; padding: 0; }}
  .printbar {{ display: none !important; }}
  .day, .block, .trow {{ break-inside: avoid; }}
}}
""".strip()


def render_html(metadata: dict, slots: list[dict]) -> str:
    meta = dict(metadata or {})
    if not meta.get("accessibilite"):
        meta["accessibilite"] = DEFAULT_ACCESSIBILITE
    if not meta.get("sanction"):
        meta["sanction"] = DEFAULT_SANCTION

    timetable = build_timetable(slots or [])

    logo = _logo_data_uri()
    logo_html = f'<img class="logo" src="{logo}" alt="GEPROMED">' if logo else \
        '<div style="font-weight:800;color:var(--navy);font-size:18pt;">GEPROMED</div>'

    ref_bits = []
    if meta.get("reference"):
        ref_bits.append(f"Réf. {_esc(meta['reference'])}")
    ref_bits.append(f"Version {_esc(meta.get('version', '[v.]'))}")
    ref_bits.append(f"Mise à jour : {_esc(meta.get('date', '[date]'))}")
    refline = "  ·  ".join(ref_bits)

    title = _esc(meta.get("intitule") or "[Intitulé de la formation]")

    # Fiche blocks (skip empty optional blocks).
    block_html = []
    for key, label, required in BLOCKS:
        value = meta.get(key)
        empty = value is None or value == "" or value == []
        if not required and empty and key in ("moyens", "sanction", "contact", "indicateurs"):
            continue
        block_html.append(
            f'<section class="block"><h2>{_esc(label)}</h2>{_render_block_value(value)}</section>'
        )

    compliance = (
        "Note de conformité : ce programme suit les exigences du Référentiel National "
        "Qualité (Qualiopi). Vérifier que chaque objectif est évaluable et couvert par "
        "les modalités d'évaluation. Toute valeur entre crochets doit être confirmée par "
        "le Responsable Qualité avant diffusion publique."
    )

    doc = f"""<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title} — Programme de formation GEPROMED</title>
<style>{CSS}</style>
</head>
<body>
<div class="printbar">
  <span>Programme de formation GEPROMED — document prêt à imprimer</span>
  <button onclick="window.print()">Imprimer / Enregistrer en PDF</button>
</div>
<div class="sheet">
  <header class="masthead">
    <div>{logo_html}<div class="brandline">GEPROMED — The medical device hub for patient safety</div></div>
    <div class="doc-tag">Programme de formation</div>
  </header>
  <h1 class="title">{title}</h1>
  <p class="refline">{refline}</p>
  {"".join(block_html)}
  {_timetable_html(timetable)}
  <div class="compliance">{_esc(compliance)}</div>
  <div class="footer">GEPROMED — organisme de formation certifié Qualiopi · ISO 9001 · ISO 13485
    &nbsp;|&nbsp; Document de travail — à valider par le Responsable Qualité avant diffusion</div>
</div>
</body>
</html>"""
    return doc


# ---------------------------------------------------------------------------
# Output writing (HTML always; PDF when WeasyPrint present)
# ---------------------------------------------------------------------------
def write_output(html_doc: str, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wants_pdf = out_path.suffix.lower() == ".pdf"

    if wants_pdf:
        try:
            from weasyprint import HTML  # type: ignore
            HTML(string=html_doc).write_pdf(str(out_path))
            return out_path
        except Exception as e:  # noqa: BLE001  (ImportError or render error)
            fallback = out_path.with_suffix(".html")
            fallback.write_text(html_doc, encoding="utf-8")
            print(
                f"NOTE: WeasyPrint unavailable ({type(e).__name__}); wrote print-ready "
                f"HTML instead. Open it and use 'Print → Save as PDF'.",
                file=sys.stderr,
            )
            return fallback

    out_path.write_text(html_doc, encoding="utf-8")
    return out_path


# ---------------------------------------------------------------------------
# Demo + sample workbook
# ---------------------------------------------------------------------------
DEMO_META = {
    "intitule": "Bootcamp Vasculaire — abord et anastomose sur simulateur",
    "reference": "GEP-FORM-VASC-01",
    "version": "1.0",
    "date": "2026-06-20",
    "public_vise": [
        "Chirurgiens vasculaires en exercice",
        "Internes en chirurgie vasculaire (à partir de la 3e année)",
    ],
    "prerequis": "Statut de praticien ou d'interne en chirurgie vasculaire. Aucun prérequis académique supplémentaire.",
    "objectifs": [
        "Réaliser une anastomose termino-latérale sur simulateur dans le temps imparti.",
        "Identifier et corriger les défauts de suture vasculaire les plus fréquents.",
        "Appliquer les principes d'exposition et de préparation du champ opératoire.",
        "Évaluer la qualité d'une anastomose à l'aide d'une grille standardisée.",
    ],
    "duree": "2 jours — 14 heures (4 demi-journées).",
    "modalites_pedagogiques": [
        "Présentiel au René Kieny Education Center (Strasbourg).",
        "Formation par simulation : ateliers pratiques sur simulateurs vasculaires (dry-lab).",
        "Pédagogie active : démonstrations, mises en situation, débriefing individualisé.",
    ],
    "moyens": [
        "Simulateurs vasculaires et consommables d'entraînement fournis.",
        "Encadrement par des chirurgiens vasculaires formateurs.",
    ],
    "evaluation": [
        "Pré-test et post-test de connaissances.",
        "Évaluation pratique sur simulateur à l'aide d'une grille standardisée.",
        "Questionnaire de satisfaction à chaud ; questionnaire à froid à [N] semaines.",
    ],
    "delais_acces": "Inscription possible jusqu'à [N] jours avant la session, dans la limite des places disponibles.",
    "tarifs": "[montant] € net de taxe par participant. Prises en charge possibles : OPCO, employeur, financement personnel.",
    "inscription": "Inscription par formulaire en ligne ou par email auprès du référent pédagogique, jusqu'au [date limite].",
    "contact": "Référent pédagogique GEPROMED — René Kieny Education Center : [email / téléphone à confirmer].",
}

DEMO_SLOTS = [
    {"jour": "Jour 1 — Lundi", "debut": "09:00", "fin": "10:30",
     "intitule": "Accueil, rappels d'anatomie chirurgicale et principes d'abord",
     "type": "Cours", "groupe": "Tous", "salle": "Amphi A", "encadrants": "Dr. Martin", "evalue": "Non"},
    {"jour": "Jour 1 — Lundi", "debut": "10:45", "fin": "12:30",
     "intitule": "Atelier suture vasculaire — dry-lab",
     "type": "Atelier pratique", "groupe": "A", "salle": "Sim-Lab 1", "encadrants": "Dr. Martin", "evalue": "Non"},
    {"jour": "Jour 1 — Lundi", "debut": "10:45", "fin": "12:30",
     "intitule": "Atelier exposition du champ opératoire",
     "type": "Atelier pratique", "groupe": "B", "salle": "Sim-Lab 2", "encadrants": "Dr. Nguyen", "evalue": "Non"},
    {"jour": "Jour 1 — Lundi", "debut": "14:00", "fin": "17:00",
     "intitule": "Anastomoses termino-latérales — mise en situation",
     "type": "Atelier pratique", "groupe": "Tous", "salle": "Sim-Lab 1+2", "encadrants": "Équipe formatrice", "evalue": "Non"},
    {"jour": "Jour 2 — Mardi", "debut": "09:00", "fin": "12:00",
     "intitule": "Débriefing individualisé et perfectionnement du geste",
     "type": "Atelier pratique", "groupe": "A", "salle": "Sim-Lab 1", "encadrants": "Dr. Martin", "evalue": "Non"},
    {"jour": "Jour 2 — Mardi", "debut": "09:00", "fin": "12:00",
     "intitule": "Débriefing individualisé et perfectionnement du geste",
     "type": "Atelier pratique", "groupe": "B", "salle": "Sim-Lab 2", "encadrants": "Dr. Nguyen", "evalue": "Non"},
    {"jour": "Jour 2 — Mardi", "debut": "13:30", "fin": "16:00",
     "intitule": "Évaluation pratique sur grille et synthèse des axes de progression",
     "type": "Cours", "groupe": "Tous", "salle": "Amphi A", "encadrants": "Équipe formatrice", "evalue": "Oui"},
]


def make_sample_xlsx(out_path: Path) -> Path:
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    # Fiche sheet
    ws = wb.active
    ws.title = "Fiche"
    ws.append(["Champ", "Valeur"])
    fiche_rows = [
        ("Intitulé", DEMO_META["intitule"]),
        ("Référence", DEMO_META["reference"]),
        ("Version", DEMO_META["version"]),
        ("Date", DEMO_META["date"]),
    ]
    for k in ("public_vise", "objectifs", "modalites_pedagogiques", "moyens", "evaluation"):
        label = {
            "public_vise": "Public visé", "objectifs": "Objectifs",
            "modalites_pedagogiques": "Modalités pédagogiques", "moyens": "Moyens",
            "evaluation": "Évaluation",
        }[k]
        for item in DEMO_META[k]:
            fiche_rows.append((label, item))
    fiche_rows += [
        ("Prérequis", DEMO_META["prerequis"]),
        ("Durée", DEMO_META["duree"]),
        ("Délais d'accès", DEMO_META["delais_acces"]),
        ("Tarifs", DEMO_META["tarifs"]),
        ("Inscription", DEMO_META["inscription"]),
        ("Contact", DEMO_META["contact"]),
    ]
    for row in fiche_rows:
        ws.append(list(row))

    # Planning sheet
    ps = wb.create_sheet("Planning")
    ps.append(SCHEDULE_COLUMNS)
    for s in DEMO_SLOTS:
        ps.append([s["jour"], s["debut"], s["fin"], s["intitule"], s["type"],
                   "" if s["groupe"] == "Tous" else s["groupe"], s["salle"],
                   s["encadrants"], s["evalue"]])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


SCHEMA_TEXT = {
    "sheets": {
        "Fiche": "Two columns 'Champ' | 'Valeur'. One Qualiopi block per row. "
                 "Repeat a Champ on several rows to build a bullet list (e.g. one "
                 "objective per row). Recognised champs: Intitulé, Référence, Version, "
                 "Date, Public visé, Prérequis, Objectifs, Durée, Modalités pédagogiques, "
                 "Moyens, Évaluation, Sanction, Accessibilité, Délais d'accès, Tarifs, "
                 "Inscription, Contact, Indicateurs.",
        "Planning": "One row per créneau. Columns (exact order): " + " | ".join(SCHEDULE_COLUMNS),
    },
    "parallel_rule": "Rows with the same Jour + Heure début + Heure fin but different "
                     "Groupe render as parallel columns. Groupe empty or 'Tous' = full-width slot.",
    "metadata_alternative": "Instead of the Fiche sheet, pass --meta fiche.json "
                            "(same keys as the internal metadata dict).",
}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(
        description="Generate a Qualiopi GEPROMED training program as print-ready HTML/PDF from an Excel schedule."
    )
    ap.add_argument("--in", dest="infile", help="Path to the schedule .xlsx (Planning sheet; optional Fiche sheet).")
    ap.add_argument("--meta", dest="metafile", help="Optional companion metadata JSON (overrides the Fiche sheet).")
    ap.add_argument("--out", default="program.html", help="Output .html or .pdf path.")
    ap.add_argument("--demo", action="store_true", help="Render the bundled demo program (no input files).")
    ap.add_argument("--make-sample-xlsx", dest="sample", help="Write a ready-to-fill sample workbook and exit.")
    ap.add_argument("--print-schema", action="store_true", help="Print the Excel schema and exit.")
    args = ap.parse_args()

    if args.print_schema:
        print(json.dumps(SCHEMA_TEXT, ensure_ascii=False, indent=2))
        return 0

    if args.sample:
        out = make_sample_xlsx(Path(args.sample))
        print(f"OK — wrote sample workbook {out}")
        return 0

    if args.demo:
        metadata, slots = DEMO_META, DEMO_SLOTS
    elif args.infile:
        try:
            metadata, slots = read_workbook(Path(args.infile))
        except SystemExit:
            raise
        except Exception as e:  # noqa: BLE001
            print(f"ERROR reading {args.infile}: {e}")
            return 2
    else:
        print("ERROR: provide --in schedule.xlsx (or --demo / --make-sample-xlsx / --print-schema).")
        return 2

    if args.metafile:
        try:
            override = json.loads(Path(args.metafile).read_text(encoding="utf-8"))
            metadata = {**(metadata or {}), **override}
        except Exception as e:  # noqa: BLE001
            print(f"ERROR reading metadata {args.metafile}: {e}")
            return 2

    html_doc = render_html(metadata, slots)
    out = write_output(html_doc, Path(args.out))
    print(f"OK — wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
